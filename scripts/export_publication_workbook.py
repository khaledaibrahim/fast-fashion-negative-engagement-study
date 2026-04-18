from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


PROJECT_ROOT = Path(__file__).resolve().parents[1]
TABLES_DIR = PROJECT_ROOT / "outputs" / "tables"
DATA_DIR = PROJECT_ROOT / "data" / "processed"
FIGURES_DIR = PROJECT_ROOT / "outputs" / "figures"


def build_brand_summary(scored_reviews: pd.DataFrame) -> pd.DataFrame:
    summary = (
        scored_reviews.groupby("brand", observed=False)
        .agg(
            n_reviews=("review_full_text", "size"),
            mean_rating=("reviewer_rating", "mean"),
            mean_sentiment=("compound", "mean"),
        )
        .reset_index()
    )

    sentiment_breakdown = (
        scored_reviews.groupby(["brand", "sentiment_label"], observed=False)
        .size()
        .reset_index(name="n_reviews")
        .pivot(index="brand", columns="sentiment_label", values="n_reviews")
        .fillna(0)
        .reset_index()
    )

    for column in ["positive", "negative", "neutral"]:
        if column not in sentiment_breakdown.columns:
            sentiment_breakdown[column] = 0

    sentiment_breakdown["total_reviews"] = (
        sentiment_breakdown["positive"] + sentiment_breakdown["negative"] + sentiment_breakdown["neutral"]
    )
    sentiment_breakdown["pct_positive"] = 100 * sentiment_breakdown["positive"] / sentiment_breakdown["total_reviews"]
    sentiment_breakdown["pct_negative"] = 100 * sentiment_breakdown["negative"] / sentiment_breakdown["total_reviews"]
    sentiment_breakdown["pct_neutral"] = 100 * sentiment_breakdown["neutral"] / sentiment_breakdown["total_reviews"]

    merged = summary.merge(
        sentiment_breakdown[["brand", "pct_positive", "pct_negative", "pct_neutral"]],
        on="brand",
        how="left",
    )
    return merged.round(
        {
            "mean_rating": 3,
            "mean_sentiment": 3,
            "pct_positive": 3,
            "pct_negative": 3,
            "pct_neutral": 3,
        }
    )


def build_topic_summary(
    topic_terms: pd.DataFrame,
    coherence: pd.DataFrame,
    construct_mapping: pd.DataFrame,
    model_name: str,
    topic_id_col: str,
) -> pd.DataFrame:
    topic_terms = topic_terms.copy()
    top_words = (
        topic_terms.sort_values([topic_id_col, "rank"])
        .groupby([topic_id_col, "publication_label"], observed=False)["term"]
        .apply(lambda s: ", ".join(s.head(10).tolist()))
        .reset_index(name="top_words")
    )

    coherence_subset = coherence.loc[(coherence["model"] == model_name) & (coherence["topic_id"] != "overall")].copy()
    coherence_subset["topic_id"] = pd.to_numeric(coherence_subset["topic_id"], errors="coerce")
    coherence_subset = coherence_subset.rename(columns={"coherence_cv": "coherence_cv_score"})

    construct_subset = construct_mapping.loc[construct_mapping["model"] == model_name].copy()

    summary = top_words.merge(
        coherence_subset[["topic_id", "coherence_cv_score"]],
        left_on=topic_id_col,
        right_on="topic_id",
        how="left",
    )
    if "topic_id_y" in summary.columns:
        summary = summary.drop(columns=["topic_id_y"]).rename(columns={"topic_id_x": "topic_id"})
    elif topic_id_col != "topic_id":
        summary = summary.rename(columns={topic_id_col: "topic_id"})

    summary = summary.merge(
        construct_subset[["publication_label", "primary_construct", "interpretation"]],
        on="publication_label",
        how="left",
    )
    return summary.sort_values("topic_id").reset_index(drop=True)


def build_readme_sheet() -> pd.DataFrame:
    rows = [
        ("Purpose", "This workbook is the publication-facing Study 1 summary based on the 1-3 star exploratory review sample only."),
        ("Scored Reviews", "Review-level export with sentiment, LDA topic, BERTopic topic, aspect marker, and emotion marker assignments."),
        ("Brand Summary", "Brand-level counts, average rating, average sentiment, and sentiment composition."),
        ("LDA Topic Summary", "Primary exploratory topic model for the paper, including publication labels, coherence, and construct interpretation."),
        ("BERTopic Summary", "Complementary semantic topic model used as a robustness and interpretive extension."),
        ("Topic Distribution", "Brand-by-topic distribution using the publication-facing LDA labels."),
        ("Aspect Distribution", "Brand-by-aspect distribution for delivery, returns, service, quality, in-store, and trust/reputation markers."),
        ("LDA Representative Quotes", "Illustrative review excerpts selected from high-membership reviews within each LDA topic."),
        ("Study1 Manuscript Table", "Compact manuscript-style table summarising the main LDA themes, coherence, construct linkage, and dominant brands."),
        ("Construct Mapping", "Mapping from Study 1 themes to LSQ, NPE, NBE, BH, and later linkage to brand switching in Study 2."),
        ("Study2 Alignment", "Direct bridge from the Study 1 exploratory outputs to the exact constructs used in Study 2."),
        ("Interpretation", "Use LDA as the main reported topic structure and BERTopic as a secondary robustness layer."),
    ]
    return pd.DataFrame(rows, columns=["sheet_or_section", "description"])


def build_representative_quotes(scored_reviews: pd.DataFrame, lda_summary: pd.DataFrame, quotes_per_topic: int = 2) -> pd.DataFrame:
    rows = []
    for topic_id, publication_label in lda_summary[["topic_id", "publication_label"]].drop_duplicates().itertuples(index=False):
        topic_col = f"topic_{int(topic_id)}"
        subset = scored_reviews.loc[scored_reviews["dominant_topic_id"] == topic_id].copy()
        if subset.empty or topic_col not in subset.columns:
            continue
        subset["membership_score"] = pd.to_numeric(subset[topic_col], errors="coerce")
        subset["quote_length_distance"] = (subset["token_count"] - 70).abs()
        subset = subset.sort_values(
            ["membership_score", "quote_length_distance", "compound"],
            ascending=[False, True, True],
        )
        seen_quotes = set()
        rank = 0
        for _, row in subset.iterrows():
            quote = str(row["review_full_text"]).strip().replace("\n", " ")
            if len(quote) < 80:
                continue
            if quote in seen_quotes:
                continue
            seen_quotes.add(quote)
            rank += 1
            rows.append(
                {
                    "topic_id": int(topic_id),
                    "publication_label": publication_label,
                    "quote_rank": rank,
                    "brand": row["brand"],
                    "reviewer_rating": row["reviewer_rating"],
                    "membership_score": round(float(row["membership_score"]), 4),
                    "compound_sentiment": round(float(row["compound"]), 4),
                    "quote_excerpt": quote[:600],
                }
            )
            if rank >= quotes_per_topic:
                break
    return pd.DataFrame(rows)


def build_manuscript_table(
    lda_summary: pd.DataFrame,
    lda_distribution: pd.DataFrame,
    aspect_distribution: pd.DataFrame,
) -> pd.DataFrame:
    def _join_brand_lines(group: pd.DataFrame) -> str:
        return "; ".join(f"{row.brand} ({row.pct_reviews:.1f}%)" for row in group.itertuples(index=False))

    def _join_aspect_lines(group: pd.DataFrame) -> str:
        return "; ".join(f"{row.aspect_top_category} ({row.pct_reviews:.1f}%)" for row in group.itertuples(index=False))

    dominant_brands = (
        lda_distribution.sort_values(["publication_label", "pct_reviews"], ascending=[True, False])
        .groupby("publication_label", observed=False)
        .head(2)
        .groupby("publication_label", observed=False)
        .agg(top_brands=("brand", lambda s: _join_brand_lines(lda_distribution.loc[s.index, ["brand", "pct_reviews"]])))
        .reset_index()
    )

    dominant_aspects = (
        aspect_distribution.sort_values(["brand", "pct_reviews"], ascending=[True, False])
        .groupby("brand", observed=False)
        .head(2)
        .groupby("brand", observed=False)
        .agg(
            top_aspects=(
                "aspect_top_category",
                lambda s: _join_aspect_lines(aspect_distribution.loc[s.index, ["aspect_top_category", "pct_reviews"]]),
            )
        )
        .reset_index()
    )

    overall_aspects = "; ".join(
        [
            f"{row.aspect_top_category} ({row.pct_reviews:.1f}%)"
            for row in aspect_distribution.groupby("aspect_top_category", observed=False)["n_reviews"].sum()
            .reset_index()
            .assign(pct_reviews=lambda d: 100 * d["n_reviews"] / d["n_reviews"].sum())
            .sort_values("pct_reviews", ascending=False)
            .head(3)
            .itertuples(index=False)
        ]
    )

    table = lda_summary.merge(dominant_brands, on="publication_label", how="left")
    table["overall_top_aspects"] = overall_aspects
    table = table[
        [
            "topic_id",
            "publication_label",
            "primary_construct",
            "coherence_cv_score",
            "top_words",
            "top_brands",
            "overall_top_aspects",
            "interpretation",
        ]
    ].rename(
        columns={
            "publication_label": "lda_theme",
            "primary_construct": "linked_construct",
            "coherence_cv_score": "coherence_cv",
        }
    )
    return table.sort_values("topic_id").reset_index(drop=True)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export a publication-style Study 1 workbook.")
    parser.add_argument(
        "--output",
        default=str(TABLES_DIR / "study1_publication_workbook.xlsx"),
        help="Path to output workbook.",
    )
    args = parser.parse_args()

    scored_reviews = pd.read_csv(DATA_DIR / "reviews_1_3_scored.csv")
    coherence = pd.read_csv(TABLES_DIR / "topic_coherence.csv")
    construct_mapping = pd.read_csv(TABLES_DIR / "construct_mapping.csv")
    lda_terms = pd.read_csv(TABLES_DIR / "topic_terms.csv")
    bertopic_terms = pd.read_csv(TABLES_DIR / "bertopic_terms.csv")
    lda_distribution = pd.read_csv(TABLES_DIR / "topic_distribution_by_brand.csv")
    aspect_distribution = pd.read_csv(TABLES_DIR / "aspect_distribution_by_brand.csv")
    study2_alignment = pd.read_csv(TABLES_DIR / "study2_alignment.csv")

    scored_export = scored_reviews[
        [
            "brand",
            "reviewer_country",
            "reviewer_published_date",
            "review_title",
            "review_text",
            "reviewer_rating",
            "compound",
            "pos",
            "neg",
            "neu",
            "sentiment_label",
            "dominant_topic_id",
            "dominant_topic_publication_label",
            "bertopic_topic_id",
            "bertopic_topic_publication_label",
            "aspect_top_category",
            "emotion_top_category",
        ]
    ].copy()
    scored_export = scored_export.rename(
        columns={
            "brand": "brand",
            "compound": "sent_compound",
            "pos": "sent_pos",
            "neg": "sent_neg",
            "neu": "sent_neu",
            "dominant_topic_id": "lda_topic",
            "dominant_topic_publication_label": "lda_topic_label",
            "bertopic_topic_id": "bertopic_topic",
            "bertopic_topic_publication_label": "bertopic_topic_label",
        }
    )

    brand_summary = build_brand_summary(scored_reviews)
    lda_summary = build_topic_summary(
        lda_terms,
        coherence,
        construct_mapping,
        model_name="LDA",
        topic_id_col="topic_id",
    )
    bertopic_summary = build_topic_summary(
        bertopic_terms.loc[bertopic_terms["topic_id"] != -1],
        coherence,
        construct_mapping,
        model_name="BERTopic",
        topic_id_col="topic_id",
    )
    lda_distribution_export = lda_distribution[
        ["brand", "publication_label", "n_reviews", "pct_reviews"]
    ].rename(columns={"publication_label": "lda_topic_label"})
    aspect_distribution_export = aspect_distribution.rename(
        columns={"aspect_top_category": "aspect_label"}
    )
    lda_quotes = build_representative_quotes(scored_reviews, lda_summary, quotes_per_topic=2)
    manuscript_table = build_manuscript_table(lda_summary, lda_distribution, aspect_distribution)
    readme_sheet = build_readme_sheet()

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        readme_sheet.to_excel(writer, sheet_name="Read Me", index=False)
        scored_export.to_excel(writer, sheet_name="Scored Reviews", index=False)
        brand_summary.to_excel(writer, sheet_name="Brand Summary", index=False)
        lda_summary.to_excel(writer, sheet_name="LDA Topic Summary", index=False)
        bertopic_summary.to_excel(writer, sheet_name="BERTopic Summary", index=False)
        lda_distribution_export.to_excel(writer, sheet_name="Topic Distribution", index=False)
        aspect_distribution_export.to_excel(writer, sheet_name="Aspect Distribution", index=False)
        lda_quotes.to_excel(writer, sheet_name="LDA Representative Quotes", index=False)
        manuscript_table.to_excel(writer, sheet_name="Study1 Manuscript Table", index=False)
        construct_mapping.to_excel(writer, sheet_name="Construct Mapping", index=False)
        study2_alignment.to_excel(writer, sheet_name="Study2 Alignment", index=False)

    workbook = load_workbook(output_path)
    figure_specs = [
        ("Figure 1", FIGURES_DIR / "fig1_sentiment_distribution.png"),
        ("Figure 2", FIGURES_DIR / "fig2_sentiment_vs_rating.png"),
        ("Figure 3", FIGURES_DIR / "fig3_lda_heatmap.png"),
        ("Figure 4", FIGURES_DIR / "fig4_topic_distribution.png"),
        ("Figure 5", FIGURES_DIR / "fig5_aspect_distribution.png"),
        ("Figure 6", FIGURES_DIR / "fig6_sentiment_trend.png"),
        ("Figure 7", FIGURES_DIR / "fig7_emotion_prevalence.png"),
        ("Figure 8", FIGURES_DIR / "fig8_topic_coherence_comparison.png"),
    ]
    for sheet_name, figure_path in figure_specs:
        if not figure_path.exists():
            continue
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            ws = workbook.create_sheet(title=sheet_name[:31])
        ws["A1"] = sheet_name
        ws["A2"] = str(figure_path)
        img = XLImage(str(figure_path))
        img.width = img.width * 0.72
        img.height = img.height * 0.72
        ws.add_image(img, "A4")
    workbook.save(output_path)

    print(f"Wrote workbook to {output_path}")


if __name__ == "__main__":
    main()
